import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Load data cleaned
df = pd.read_csv("hotel_bookings_cleaned.csv")

# Folder output
output_excel = "visualization_results.xlsx"

# ================================
# 1. Buat grafik dan simpan PNG
# ================================

# ---- BAR CHART
plt.figure(figsize=(10,5))
df['hotel'].value_counts().plot(kind='bar')
plt.title("Jumlah Booking Berdasarkan Tipe Hotel")
plt.xlabel("Tipe Hotel")
plt.ylabel("Jumlah Booking")
plt.tight_layout()
bar_path = "bar_chart.png"
plt.savefig(bar_path)
plt.close()

# ---- LINE CHART
monthly_booking = df.groupby('arrival_date_month').size().reindex([
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
])

plt.figure(figsize=(12,5))
monthly_booking.plot(kind='line', marker='o')
plt.title("Tren Jumlah Booking Per Bulan")
plt.xlabel("Bulan")
plt.ylabel("Jumlah Booking")
plt.grid(True)
plt.tight_layout()
line_path = "line_chart.png"
plt.savefig(line_path)
plt.close()

# ---- PIE CHART
plt.figure(figsize=(7,7))
df['customer_type'].value_counts().plot(kind='pie', autopct='%1.1f%%')
plt.title("Distribusi Tipe Customer")
plt.ylabel("")
pie_path = "pie_chart.png"
plt.savefig(pie_path)
plt.close()

# ---- HEATMAP
plt.figure(figsize=(12,8))
sns.heatmap(df.corr(numeric_only=True), annot=True, cmap="coolwarm")
plt.title("Heatmap Korelasi Antar Variabel Numerik")
heatmap_path = "heatmap.png"
plt.savefig(heatmap_path)
plt.close()

# ---- SCATTER
plt.figure(figsize=(10,6))
plt.scatter(df['lead_time'], df['adr'], alpha=0.4)
plt.title("Hubungan Lead Time dan Harga (ADR)")
plt.xlabel("Lead Time")
plt.ylabel("Average Daily Rate (ADR)")
scatter_path = "scatter_plot.png"
plt.savefig(scatter_path)
plt.close()

# ================================
# 2. Masukkan semua grafik ke Excel
# ================================

wb = Workbook()
ws1 = wb.active
ws1.title = "Bar Chart"

# Tambahkan tiap gambar
image_files = {
    "Bar Chart": bar_path,
    "Line Chart": line_path,
    "Pie Chart": pie_path,
    "Heatmap": heatmap_path,
    "Scatter Plot": scatter_path
}

# Loop masukkan gambar ke tiap sheet
for sheet_name, img_path in image_files.items():
    ws = wb.create_sheet(title=sheet_name)
    img = Image(img_path)
    img.width = 900
    img.height = 500
    ws.add_image(img, "A1")

# Hapus sheet default jika kosong
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(output_excel)

print("SEMUA VISUALISASI BERHASIL DISIMPAN ke:", output_excel)

