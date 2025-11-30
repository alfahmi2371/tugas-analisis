import streamlit as st
import pandas as pd
import plotly.express as px
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# ============= LOAD DATA =============
df = pd.read_csv("hotel_bookings_cleaned.csv")

month_order = [
    "January","February","March","April","May","June",
    "July","August","September","October","November","December"
]

# ============= SIDEBAR FILTERS =============
st.sidebar.title("🔎 Filter Data")

hotel_filter = st.sidebar.multiselect(
    "Pilih Tipe Hotel",
    options=df['hotel'].unique(),
    default=df['hotel'].unique()
)

country_filter = st.sidebar.multiselect(
    "Pilih Negara",
    options=df['country'].unique(),
    default=df['country'].unique()
)

month_filter = st.sidebar.multiselect(
    "Pilih Bulan",
    options=month_order,
    default=month_order
)

df_filtered = df[
    (df['hotel'].isin(hotel_filter)) &
    (df['country'].isin(country_filter)) &
    (df['arrival_date_month'].isin(month_filter))
]

# ============= HEADER + LOGO =============
logo_path = "logo.png"  # Pastikan file logo bernama logo.png
st.image(logo_path, width=180)

st.title("🏨 Dashboard Hotel Booking Demand")
st.write("Dashboard interaktif analisis pemesanan hotel menggunakan Streamlit.")

st.markdown("---")

# ============= DATA TABLE =============
st.subheader("📄 Data Table")
st.dataframe(df_filtered, use_container_width=True)

st.markdown("---")

# ============= STATISTIK DESKRIPTIF =============
st.subheader("📊 Statistik Data (Mean – Median – Modus – Std Dev)")

if df_filtered.empty:
    st.warning("⚠ Data kosong karena filter. Silakan ubah filter agar statistik tersedia.")
else:
    numeric_df = df_filtered.select_dtypes(include=['int64', 'float64', 'number'])
    if numeric_df.empty:
        st.info("ℹ Tidak ada kolom numerik untuk dihitung statistik.")
    else:
        stats = pd.DataFrame({
            "Mean": numeric_df.mean(),
            "Median": numeric_df.median(),
            "Modus": numeric_df.mode().iloc[0],
            "Std Dev": numeric_df.std()
        }).round(2)

        stats.reset_index(inplace=True)
        stats.rename(columns={"index": "Variabel"}, inplace=True)

        st.markdown("""
            <style>
                .stats-card {
                    background-color: #f9fbff;
                    border: 1px solid #c9d9ff;
                    padding: 16px;
                    border-radius: 12px;
                    box-shadow: 0 3px 7px rgba(0,0,0,0.08);
                }
            </style>
        """, unsafe_allow_html=True)

        st.markdown("<div class='stats-card'>", unsafe_allow_html=True)
        st.dataframe(stats, use_container_width=True)
        st.markdown("</div>", unsafe_allow_html=True)

st.markdown("---")

# ============= VISUALISASI DATA =============

# Chart 1 - Bar
st.subheader("📊 Bar Chart — Jumlah Booking per Hotel")
bar_data = df_filtered['hotel'].value_counts().reset_index()
bar_data.columns = ['hotel_type', 'count']
bar_fig = px.bar(bar_data, x='hotel_type', y='count', color='hotel_type')
st.plotly_chart(bar_fig)
bar_fig.write_image("bar_chart.png")

# Chart 2 - Line
st.subheader("📈 Line Chart — Tren Booking per Bulan")
monthly_booking = df_filtered.groupby("arrival_date_month").size().reindex(month_order)
line_fig = px.line(x=month_order, y=monthly_booking.values, markers=True)
st.plotly_chart(line_fig)
line_fig.write_image("line_chart.png")

# Chart 3 - Pie Chart
st.subheader("🥧 Pie Chart — Customer Type")
pie_fig = px.pie(df_filtered, names='customer_type')
st.plotly_chart(pie_fig)
pie_fig.write_image("pie_chart.png")

# Chart 4 - Scatter
st.subheader("📍 Scatter Plot — Lead Time vs ADR")
scatter_fig = px.scatter(df_filtered, x='lead_time', y='adr', color='hotel')
st.plotly_chart(scatter_fig)
scatter_fig.write_image("scatter_plot.png")

# Chart 5 - Heatmap
st.subheader("🔥 Heatmap — Korelasi Variabel Numerik")
plt.figure(figsize=(8,5))
sns.heatmap(df_filtered.select_dtypes(include="number").corr(), annot=True, cmap="coolwarm")
plt.tight_layout()
plt.savefig("heatmap.png")
plt.close()
st.image("heatmap.png")

st.markdown("---")

# ============= EXPORT EXCEL =============
st.subheader("⬇ Download Semua Data & Grafik sebagai Excel")

def export_to_excel():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Filtered Data"
    for i, col in enumerate(df_filtered.columns, 1):
        ws1.cell(row=1, column=i).value = col
    for r, row in enumerate(df_filtered.values, 2):
        for c, v in enumerate(row, 1):
            ws1.cell(row=r, column=c).value = v

    image_files = {
        "Bar Chart": "bar_chart.png",
        "Line Chart": "line_chart.png",
        "Pie Chart": "pie_chart.png",
        "Scatter Plot": "scatter_plot.png",
        "Heatmap": "heatmap.png"
    }

    for sheet, img_path in image_files.items():
        ws = wb.create_sheet(title=sheet)
        img = Image(img_path)
        img.width = 900
        img.height = 500
        ws.add_image(img, "A1")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

excel_file = export_to_excel()

st.download_button(
    label="📥 Download dashboard_export.xlsx",
    data=excel_file,
    file_name="dashboard_export.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

st.markdown("---")
st.write("Dashboard dibuat untuk Tugas UAS Analisis & Visualisasi Data 🎓")
